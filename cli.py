"""Non-interactive batch interface for SEM Ready."""

from __future__ import annotations

import argparse
import csv
from datetime import datetime, timezone
import glob
import hashlib
import json
from pathlib import Path
import re
import sys
import time

from sem_ready import (EnhancementOptions, SUPPORTED_EXTENSIONS, analyze, export,
                       create_qc_sheet, nice_scale_value, PUBLICATION_PROFILES)


REPORT_FIELDS = [
    "status", "source", "output", "crop_y", "crop_confidence", "hfw_um",
    "hfw_confidence", "scale_um", "scale_confidence", "bar_pixels",
    "effective_dpi", "output_bytes", "vendor", "profile", "source_sha256",
    "elapsed_seconds", "warnings", "error",
]


def parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Automatically crop SEM footers and add HFW-calibrated scale bars.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("inputs", nargs="*", help="Images, directories, or wildcard patterns")
    p.add_argument("--config", type=Path, help="JSON settings profile")
    p.add_argument("-o", "--output", type=Path, default=Path("publication_ready"))
    p.add_argument("--recursive", action="store_true", help="Search input directories recursively")
    p.add_argument("--overrides-csv", type=Path,
                   help="Per-image values: source/filename,hfw_um,scale_um,crop_y")
    p.add_argument("--hfw-um", type=float, help="Global HFW override")
    p.add_argument("--scale-um", type=float, help="Global scale-value override")
    p.add_argument("--crop-y", type=int, help="Global footer-boundary override")
    # Accepted for compatibility with early scripts; annotations now use the
    # fixed publication style (black on a translucent white panel).
    p.add_argument("--color", choices=("white", "black"), default="black",
                   help=argparse.SUPPRESS)
    p.add_argument("--dpi", type=int, default=600)
    p.add_argument("--format", choices=("tif", "png", "jpg"), default="tif")
    p.add_argument("--name-template", default="{stem}_publication",
                   help="Output name using {stem}, {profile}, and {format}")
    p.add_argument("--profile", choices=tuple(PUBLICATION_PROFILES), default="original",
                   help="Print-size/resolution optimization profile")
    p.add_argument("--max-file-mb", type=float, help="JPEG size target in MB")
    p.add_argument("--jpeg-quality", type=int, default=95)
    p.add_argument("--strict-dpi", action="store_true",
                   help="Fail when the source cannot meet the profile DPI")
    p.add_argument("--enhancement",
                   choices=("raw", "auto-contrast", "balanced", "local-contrast",
                            "detail", "uneven-background", "inverted"), default="raw")
    p.add_argument("--scale-position",
                   choices=("auto", "bottom-right", "bottom-left", "top-right", "top-left"),
                   default="bottom-right")
    p.add_argument("--journal-preset", choices=("quarter-a4", "single-column", "double-column"),
                   default="quarter-a4")
    p.add_argument("--auto-scale", action="store_true",
                   help="Replace the OCR bar value with a rounded 1/2/5 scale")
    p.add_argument("--min-ocr-confidence", type=float, default=0.20,
                   help="Reject lower-confidence automatic HFW/scale reads")
    p.add_argument("--no-ocr", action="store_true",
                   help="Never run OCR; calibration must be supplied by overrides")
    p.add_argument("--overwrite", action="store_true", help="Replace existing outputs")
    p.add_argument("--manifest", type=Path,
                   help="Resume/deduplication manifest (default: OUTPUT/.sem_ready_manifest.json)")
    p.add_argument("--retry-failed", action="store_true",
                   help="Process only items marked failed in the existing manifest")
    p.add_argument("--fail-fast", action="store_true", help="Stop after the first failed image")
    p.add_argument("--report", type=Path,
                   help="CSV report path (default: OUTPUT/batch_report.csv)")
    p.add_argument("--no-report", action="store_true")
    p.add_argument("--jsonl", action="store_true",
                   help="Write one machine-readable JSON object per image to stdout")
    p.add_argument("--no-audit", action="store_true",
                   help="Do not create a calibration JSON sidecar for each output")
    return p


def discover_inputs(specs: list[str], recursive: bool = False) -> tuple[list[Path], list[str]]:
    """Expand files, directories, and globs consistently on Windows and Unix."""
    files: list[Path] = []
    problems: list[str] = []
    for spec in specs:
        candidate = Path(spec)
        matches: list[Path]
        if candidate.exists():
            if candidate.is_dir():
                iterator = candidate.rglob("*") if recursive else candidate.iterdir()
                matches = [p for p in iterator if p.is_file()]
            else:
                matches = [candidate]
        else:
            matches = [Path(p) for p in glob.glob(spec, recursive=recursive)]
            if not matches:
                problems.append(f"Input did not match anything: {spec}")
                continue
        supported = [p.resolve() for p in matches if p.suffix.lower() in SUPPORTED_EXTENSIONS]
        if not supported and candidate.exists():
            problems.append(f"No supported SEM images found in: {spec}")
        files.extend(supported)
    return sorted(dict.fromkeys(files), key=lambda p: str(p).lower()), problems


def load_overrides(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None:
        return {}
    if not path.exists():
        raise ValueError(f"Overrides CSV does not exist: {path}")
    lookup: dict[str, dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = set(reader.fieldnames or [])
        if not ({"source", "filename"} & headers):
            raise ValueError("Overrides CSV needs a 'source' or 'filename' column.")
        for line, row in enumerate(reader, start=2):
            key = (row.get("source") or row.get("filename") or "").strip()
            if not key:
                raise ValueError(f"Overrides CSV row {line} has no source/filename.")
            lookup[key.lower()] = row
            source_path = Path(key)
            if source_path.is_absolute():
                lookup[str(source_path.resolve()).lower()] = row
    return lookup


def override_for(source: Path, overrides: dict[str, dict[str, str]]) -> dict[str, str]:
    return (overrides.get(str(source.resolve()).lower())
            or overrides.get(source.name.lower())
            or overrides.get(source.stem.lower())
            or {})


def _number(row: dict[str, str], name: str, fallback, cast):
    raw = (row.get(name) or "").strip()
    return cast(raw) if raw else fallback


def _emit(record: dict, jsonl: bool):
    if jsonl:
        print(json.dumps(record, ensure_ascii=False), flush=True)
    elif record["status"] == "ok":
        print(f"OK       {Path(record['source']).name} -> {record['output']}")
    elif record["status"] == "skipped":
        print(f"SKIPPED  {Path(record['source']).name} (output exists)")
    else:
        print(f"ERROR    {record['source']}: {record['error']}", file=sys.stderr, flush=True)


def _enhancement(name: str) -> EnhancementOptions:
    return {
        "raw": EnhancementOptions(preserve_raw=True),
        "auto-contrast": EnhancementOptions(
            preserve_raw=False, auto_contrast=True, auto_brightness=True),
        "balanced": EnhancementOptions(
            preserve_raw=False, auto_contrast=True, auto_brightness=True, denoise=True),
        "local-contrast": EnhancementOptions(preserve_raw=False, clahe=True),
        "detail": EnhancementOptions(
            preserve_raw=False, auto_contrast=True, denoise=True, sharpen=True),
        "uneven-background": EnhancementOptions(
            preserve_raw=False, shading_correction=True, auto_contrast=True),
        "inverted": EnhancementOptions(preserve_raw=False, invert=True, auto_contrast=True),
    }[name]


def _journal(name: str) -> tuple[float, float]:
    return {"quarter-a4": (14.0, 105.0), "single-column": (10.0, 85.0),
            "double-column": (10.0, 178.0)}[name]


def _destination(source: Path, output: Path, extension: str,
                 used: dict[str, Path], template: str = "{stem}_publication",
                 profile: str = "original") -> Path:
    try:
        name = template.format(stem=source.stem, profile=profile, format=extension)
    except (KeyError, ValueError) as exc:
        raise ValueError(f"Invalid --name-template: {exc}") from exc
    name = re.sub(r'[<>:"/\\|?*]', "_", name).strip(" .")
    if not name:
        raise ValueError("Output name template produced an empty name")
    base = output / f"{name}.{extension}"
    key = str(base).lower()
    if key in used and used[key] != source:
        # Stable suffix avoids collisions when recursive folders contain the same name.
        import hashlib
        suffix = hashlib.sha1(str(source).encode("utf-8")).hexdigest()[:8]
        base = output / f"{name}_{suffix}.{extension}"
    used[str(base).lower()] = source
    return base


def write_report(path: Path, records: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REPORT_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
    temporary.replace(path)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SEM batch report"
        sheet.append(REPORT_FIELDS)
        for record in records:
            sheet.append([record.get(field, "") for field in REPORT_FIELDS])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(60, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
        excel_path = path.with_suffix(".xlsx")
        temporary_excel = excel_path.with_suffix(".xlsx.tmp")
        workbook.save(temporary_excel)
        temporary_excel.replace(excel_path)
    except ImportError:
        pass


def _parse_args(argv=None):
    preliminary = argparse.ArgumentParser(add_help=False)
    preliminary.add_argument("--config", type=Path)
    known, _ = preliminary.parse_known_args(argv)
    defaults = {}
    config_inputs = []
    if known.config:
        try:
            data = json.loads(known.config.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            preliminary.error(f"Cannot read config: {exc}")
        if not isinstance(data, dict):
            preliminary.error("Config root must be a JSON object")
        config_inputs = data.pop("inputs", [])
        defaults = data
    command_parser = parser()
    if defaults:
        valid = {action.dest for action in command_parser._actions}
        unknown = sorted(set(defaults) - valid)
        if unknown:
            command_parser.error(f"Unknown config keys: {', '.join(unknown)}")
        command_parser.set_defaults(**defaults)
    args = command_parser.parse_args(argv)
    if not args.inputs:
        args.inputs = [str(value) for value in config_inputs]
    if not args.inputs:
        command_parser.error("provide input images/folders or set 'inputs' in --config")
    return args


def _load_manifest(path: Path) -> dict:
    if not path.exists():
        return {"version": 1, "items": {}}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict) and isinstance(data.get("items"), dict):
            return data
    except (OSError, json.JSONDecodeError):
        pass
    return {"version": 1, "items": {}}


def _write_manifest(path: Path, manifest: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    temporary.replace(path)


def main(argv=None) -> int:
    args = _parse_args(argv)
    args.output = Path(args.output)
    args.overrides_csv = Path(args.overrides_csv) if args.overrides_csv else None
    args.report = Path(args.report) if args.report else None
    args.manifest = Path(args.manifest) if args.manifest else None
    if not (0 <= args.min_ocr_confidence <= 1):
        parser().error("--min-ocr-confidence must be between 0 and 1")
    if not (72 <= args.dpi <= 2400):
        parser().error("--dpi must be between 72 and 2400")
    if not (70 <= args.jpeg_quality <= 100):
        parser().error("--jpeg-quality must be between 70 and 100")
    if args.max_file_mb is not None and args.max_file_mb <= 0:
        parser().error("--max-file-mb must be positive")

    try:
        sources, discovery_problems = discover_inputs(args.inputs, args.recursive)
        overrides = load_overrides(args.overrides_csv)
    except ValueError as exc:
        print(f"CONFIG ERROR: {exc}", file=sys.stderr)
        return 2
    if discovery_problems:
        for problem in discovery_problems:
            print(f"INPUT WARNING: {problem}", file=sys.stderr)
    if not sources:
        print("No supported input images were found.", file=sys.stderr)
        return 2

    output_root = args.output.resolve()
    # A common scheduled-job layout places OUTPUT below INPUT. Never feed prior
    # generated images back into a later recursive run.
    sources = [source for source in sources if not source.is_relative_to(output_root)]
    if not sources:
        print("No source images remain after excluding the output directory.", file=sys.stderr)
        return 2
    args.output.mkdir(parents=True, exist_ok=True)
    report_path = args.report or (args.output / "batch_report.csv")
    manifest_path = args.manifest or (args.output / ".sem_ready_manifest.json")
    manifest = _load_manifest(manifest_path)
    records: list[dict] = []
    used_destinations: dict[str, Path] = {}
    failures = 0
    enhancements = _enhancement(args.enhancement)
    label_pt, figure_width_mm = _journal(args.journal_preset)
    settings_payload = {
        key: getattr(args, key) for key in (
            "format", "profile", "name_template", "dpi", "enhancement", "scale_position",
            "journal_preset", "auto_scale", "max_file_mb", "jpeg_quality", "strict_dpi")
    }
    settings_sha = hashlib.sha256(
        json.dumps(settings_payload, sort_keys=True).encode("utf-8")).hexdigest()

    for source in sources:
        started = time.perf_counter()
        destination = _destination(source, output_root, args.format, used_destinations,
                                   args.name_template, args.profile)
        record = {name: "" for name in REPORT_FIELDS}
        record.update(status="error", source=str(source), output=str(destination), profile=args.profile)
        try:
            source_sha = hashlib.sha256(source.read_bytes()).hexdigest()
            record["source_sha256"] = source_sha
            prior = manifest["items"].get(str(source))
            if args.retry_failed and (not prior or prior.get("status") != "error"):
                record.update(status="skipped", warnings="Not previously marked failed")
                continue
            duplicate = next((item for item in manifest["items"].values()
                              if item.get("source_sha256") == source_sha
                              and item.get("settings_sha256") == settings_sha
                              and item.get("status") == "ok"
                              and Path(item.get("output", "")).exists()), None)
            if duplicate and not args.overwrite:
                record.update(status="skipped", warnings=f"Duplicate of {duplicate.get('source')}")
                continue
            row = override_for(source, overrides)
            crop = _number(row, "crop_y", args.crop_y, int)
            supplied_hfw = _number(row, "hfw_um", args.hfw_um, float)
            supplied_scale = _number(row, "scale_um", args.scale_um, float)
            if destination.exists() and not args.overwrite:
                record["status"] = "skipped"
                continue
            if args.no_ocr and (supplied_hfw is None or supplied_scale is None):
                raise ValueError("--no-ocr requires HFW and scale overrides for every image.")

            # Avoid the relatively expensive OCR pass when both calibration values
            # have already been supplied by an automated acquisition workflow.
            run_ocr = not args.no_ocr and (supplied_hfw is None or supplied_scale is None)
            image, result = analyze(source, run_ocr=run_ocr, crop_y=crop)
            hfw = supplied_hfw if supplied_hfw is not None else result.hfw_um
            scale = supplied_scale if supplied_scale is not None else result.scale_um
            if hfw is None or (scale is None and not args.auto_scale):
                raise ValueError("OCR could not identify HFW and scale; add CSV overrides.")
            if supplied_hfw is None and result.hfw_confidence < args.min_ocr_confidence:
                raise ValueError(f"HFW OCR confidence {result.hfw_confidence:.2f} is below threshold.")
            if (not args.auto_scale and supplied_scale is None
                    and result.scale_confidence < args.min_ocr_confidence):
                raise ValueError(f"Scale OCR confidence {result.scale_confidence:.2f} is below threshold.")

            if args.auto_scale:
                scale = nice_scale_value(hfw)

            export(image, destination, result, hfw, scale, "black", args.dpi, not args.no_audit,
                   enhancements=enhancements, position=args.scale_position,
                   label_pt=label_pt, figure_width_mm=figure_width_mm,
                   profile=args.profile, max_file_mb=args.max_file_mb,
                   jpeg_quality=args.jpeg_quality, strict_dpi=args.strict_dpi)
            audit_path = destination.with_suffix(destination.suffix + ".json")
            audit_data = (json.loads(audit_path.read_text(encoding="utf-8"))
                          if audit_path.exists() else {})
            record.update(
                status="ok", crop_y=result.crop_y,
                crop_confidence=f"{result.crop_confidence:.3f}", hfw_um=f"{hfw:g}",
                hfw_confidence=("manual" if supplied_hfw is not None else f"{result.hfw_confidence:.3f}"),
                scale_um=f"{scale:g}",
                scale_confidence=("manual" if supplied_scale is not None else f"{result.scale_confidence:.3f}"),
                bar_pixels=round(image.width * scale / hfw),
                effective_dpi=audit_data.get("effective_dpi", ""),
                output_bytes=destination.stat().st_size,
                vendor=result.vendor_hint or "",
                warnings=" | ".join(result.warnings + [
                    f"Enhancement preset: {args.enhancement}",
                    f"Scale position: {args.scale_position}",
                ]),
            )
        except Exception as exc:
            failures += 1
            record["error"] = str(exc)
        finally:
            record["elapsed_seconds"] = f"{time.perf_counter() - started:.3f}"
            records.append(record)
            manifest["items"][str(source)] = {
                "source": str(source), "source_sha256": record.get("source_sha256", ""),
                "settings_sha256": settings_sha, "status": record["status"],
                "output": record["output"], "error": record.get("error", ""),
                "updated_utc": datetime.now(timezone.utc).isoformat(),
            }
            _write_manifest(manifest_path, manifest)
            _emit(record, args.jsonl)
            if not args.no_report:
                write_report(report_path, records)
        if failures and args.fail_fast:
            break

    if records:
        create_qc_sheet(records, args.output / "batch_qc.jpg")
        summary = {
            "tool": "SEM Ready",
            "completed_utc": datetime.now(timezone.utc).isoformat(),
            "settings": settings_payload,
            "settings_sha256": settings_sha,
            "counts": {
                "ok": sum(r["status"] == "ok" for r in records),
                "skipped": sum(r["status"] == "skipped" for r in records),
                "failed": sum(r["status"] == "error" for r in records),
            },
            "report": str(report_path.resolve()),
            "qc_sheet": str((args.output / "batch_qc.jpg").resolve()),
            "manifest": str(manifest_path.resolve()),
        }
        (args.output / "project_summary.json").write_text(
            json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    if not args.jsonl:
        ok = sum(r["status"] == "ok" for r in records)
        skipped = sum(r["status"] == "skipped" for r in records)
        print(f"\nBatch complete: {ok} exported, {skipped} skipped, {failures} failed.")
        if not args.no_report:
            print(f"Report: {report_path.resolve()}")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
